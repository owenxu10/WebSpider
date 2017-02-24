# -*- coding: utf-8 -*-

import sys
from openpyxl import load_workbook
from config import excelConfig
from elasticsearch import Elasticsearch
from openpyxl.styles import Border, Side, PatternFill

def style_range(ws):
	border = Border(left=Side(style='thin', color='FF99C4E6'),
					right=Side(style='thin', color='FF99C4E6'),
					top=Side(style='thin', color='FF99C4E6'),
					bottom=Side(style='thin', color='FF99C4E6'))

	fill = PatternFill("solid", fgColor="f9fbfd")
	for row in range(1, ws.max_row+1):
		for col in range(1, ws.max_column+1):
			ws.cell(column=col, row=row).border = border
			ws.cell(column=col, row=row).fill = fill

es = Elasticsearch()
default_encoding="utf-8"
if(default_encoding!=sys.getdefaultencoding()):
    reload(sys)
    sys.setdefaultencoding(default_encoding)

filepath = excelConfig['path'] + excelConfig['name']
wb = load_workbook(filepath)

sheet_names = wb.get_sheet_names()
sheet_num = len(sheet_names)

for sheet in range(0,sheet_num):

	ws = wb[sheet_names[sheet]]

	ws.cell(row=1, column=2).value = "结果";

	row_num = ws.max_row + 1
	for i in range(2,row_num):

		question =  ws.cell(row=i, column=1).value
		if question is not None:
			print question
			try:
				val = int(question)
			except ValueError:
				test_data = {
							  "query": {
								"bool": {
								  "should": [
									{
									  "multi_match": {
										"query": question,
										"fields": [
										  "question^2",
										  "answer^1"
										],
										"boost": 1
									  }
									},
									{
									  "match_phrase": {
										"question": {
										  "query": question,
										  "boost": 3
										}
									  }
									},
									{
									  "match_phrase": {
										"answer": {
										  "query": question,
										  "slop": 0,
										  "boost": 2
										}
									  }
									}
								  ]
								}
							  },
							  "highlight": {
								"pre_tags": [
								  ""
								],
								"post_tags": [
								  ""
								],
								"fields": {
								  "question": {},
								  "answer": {}
								}
							  }
							}

				res = es.search(index="qa_demo", body=test_data)
				maxScore=res['hits']['max_score']
				all_hits = res['hits']['hits']
				filtered_hits = filter(lambda x: x['_score']>20 and x['_score']>0.1*maxScore, all_hits)


				if(len(filtered_hits) >= 5):
					content = content = excelConfig['exist'] +"\n\n"\
							  +filtered_hits[0]['_source']['question']+'\n'\
							  + filtered_hits[1]['_source']['question']+'\n'\
							  + filtered_hits[2]['_source']['question']+'\n'\
							  + filtered_hits[3]['_source']['question']+"\n"\
							  + filtered_hits[4]['_source']['question']+"\n"
					ws.cell(row=i,column=2).value = content
				elif(len(filtered_hits) > 0):
					content = excelConfig['exist'] +"\n\n"
					for filtered_hit in filtered_hits:
						content = content + filtered_hit['_source']['question'] + "\n"
					ws.cell(row=i, column=2).value = content
				else:
					print "Not found"
					ws.cell(row=i,column=2).value= excelConfig['not_found']

style_range(ws)
wb.save(filepath)
